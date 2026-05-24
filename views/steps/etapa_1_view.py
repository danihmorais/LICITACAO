import json
import re
import customtkinter as ctk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation


def parse_moeda(texto: str) -> Decimal:
    if not texto:
        return Decimal("0")
    texto = str(texto).strip()
    texto = re.sub(r"[^\d,.-]", "", texto)
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation:
        return Decimal("0")


def formatar_brl(valor: Decimal) -> str:
    valor = valor.quantize(Decimal("0.01"))
    s = f"{valor:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


class Etapa1View(ctk.CTkFrame):
    SECTION_TITLE_COLOR = ("gray15", "gray85")
    SECTION_HELP_COLOR = ("gray45", "gray65")
    ACCENT_COLOR = "#2563EB"

    def __init__(self, master, controller):
        super().__init__(master, fg_color="transparent")

        self.controller = controller
        self.linhas_itens = []

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.container = ctk.CTkFrame(
            self,
            corner_radius=16,
            fg_color=("white", "#111827"),
            border_width=1,
            border_color=("gray85", "#1F2937"),
        )
        self.container.grid(row=0, column=0, sticky="nsew")
        self.container.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.container,
            text="Objeto da Licitação:",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=self.SECTION_TITLE_COLOR,
        ).grid(row=0, column=0, sticky="w", padx=35, pady=(25, 8))

        ctk.CTkLabel(
            self.container,
            text="Descreva brevemente o objeto licitado para direcionar a geração de especificações.",
            font=ctk.CTkFont(size=13),
            text_color=self.SECTION_HELP_COLOR,
            wraplength=680,
            justify="left",
        ).grid(row=1, column=0, sticky="w", padx=35, pady=(0, 8))

        self.entry_objeto = ctk.CTkEntry(
            self.container,
            height=42,
            corner_radius=12,
            border_width=1,
            border_color=("gray85", "#1F2937"),
            font=ctk.CTkFont(size=13),
        )
        self.entry_objeto.grid(row=2, column=0, sticky="ew", padx=35, pady=(0, 15))

        ctk.CTkLabel(
            self.container,
            text="Justificativa da Demanda:",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=self.SECTION_TITLE_COLOR,
        ).grid(row=3, column=0, sticky="w", padx=35, pady=(0, 5))

        ctk.CTkLabel(
            self.container,
            text="Descreva brevemente a justificativa da demanda.",
            font=ctk.CTkFont(size=13),
            text_color=self.SECTION_HELP_COLOR,
            wraplength=680,
            justify="left",
        ).grid(row=4, column=0, sticky="w", padx=35, pady=(0, 8))

        self.text_necessidade = ctk.CTkTextbox(
            self.container,
            height=120,
            corner_radius=12,
            border_width=1,
            border_color=("gray85", "#1F2937"),
            font=ctk.CTkFont(size=14),
        )
        self.text_necessidade.grid(row=5, column=0, sticky="ew", padx=35, pady=(0, 15))

        self.frame_tabela_container = ctk.CTkFrame(self.container, fg_color="transparent")
        self.frame_tabela_container.grid(row=6, column=0, sticky="ew", padx=35, pady=(0, 20))
        self.frame_tabela_container.grid_columnconfigure(0, weight=1)

        self.frame_topo_botoes = ctk.CTkFrame(self.frame_tabela_container, fg_color="transparent")
        self.frame_topo_botoes.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        self.btn_add_item = ctk.CTkButton(
            self.frame_topo_botoes,
            text="+ Novo Item",
            width=140,
            height=38,
            corner_radius=12,
            fg_color=self.ACCENT_COLOR,
            hover_color="#1D4ED8",
            command=self.adicionar_item_linha,
        )
        self.btn_add_item.pack(side="left", padx=(0, 10))

        self.btn_importar = ctk.CTkButton(
            self.frame_topo_botoes,
            text="Importar XLSX",
            width=140,
            height=38,
            corner_radius=12,
            fg_color=self.ACCENT_COLOR,
            hover_color="#1D4ED8",
            command=self.importar_planilha,
        )
        self.btn_importar.pack(side="left")

        header = ctk.CTkFrame(self.frame_tabela_container, fg_color=("gray75", "gray20"), corner_radius=8)
        header.grid(row=1, column=0, sticky="ew", pady=(0, 5))

        colunas = [
            ("#", 40),
            ("Descrição", 340),
            ("UN", 60),
            ("Qtd", 70),
            ("Vlr Unit.", 100),
            ("Total", 120),
            ("", 120),
        ]

        for i, (txt, width) in enumerate(colunas):
            ctk.CTkLabel(header, text=txt, width=width, font=ctk.CTkFont(weight="bold")).grid(
                row=0, column=i, padx=3, pady=8
            )

        self.scroll_itens = ctk.CTkScrollableFrame(
            self.frame_tabela_container,
            height=250,
            fg_color=("gray92", "gray14"),
            scrollbar_button_color=("#D1D5DB", "#4B5563"),
            scrollbar_button_hover_color=("#9CA3AF", "#6B7280"),
            scrollbar_fg_color="transparent",
        )
        self.scroll_itens.grid(row=2, column=0, sticky="ew")

        self.label_total_geral = ctk.CTkLabel(
            self.frame_tabela_container,
            text="TOTAL GERAL: R$ 0,00",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=self.SECTION_TITLE_COLOR,
        )
        self.label_total_geral.grid(row=3, column=0, sticky="e", pady=(10, 0))

        self.adicionar_item_linha()

    def mascara_real_time(self, entry):
        texto = entry.get()
        numeros = "".join(c for c in texto if c.isdigit())

        if not numeros:
            self.atualizar_totais()
            return

        valor = Decimal(numeros) / Decimal("100")
        novo = formatar_brl(valor)

        if novo != texto:
            entry.delete(0, "end")
            entry.insert(0, novo)
            entry.icursor(len(novo))

        self.atualizar_totais()

    def adicionar_item_linha(self):
        linha = ctk.CTkFrame(self.scroll_itens, fg_color=("gray88", "gray18"))
        linha.pack(fill="x", pady=3, padx=2)

        lbl_item = ctk.CTkLabel(linha, text="1", width=40)
        lbl_item.grid(row=0, column=0, padx=3, pady=5)

        ent_desc = ctk.CTkEntry(linha, width=340, height=38, corner_radius=10, border_width=1)
        ent_desc.grid(row=0, column=1, padx=3)

        ent_un = ctk.CTkEntry(linha, width=60, height=38, corner_radius=10, border_width=1)
        ent_un.insert(0, "UN")
        ent_un.grid(row=0, column=2, padx=3)

        ent_qtd = ctk.CTkEntry(linha, width=70, height=38, corner_radius=10, border_width=1)
        ent_qtd.insert(0, "1")
        ent_qtd.grid(row=0, column=3, padx=3)

        ent_unit = ctk.CTkEntry(linha, height=38, corner_radius=10, border_width=1)
        ent_unit.insert(0, "0,00")
        ent_unit.grid(row=0, column=4, padx=3)

        lbl_total = ctk.CTkLabel(linha, text="R$ 0,00", width=120, font=ctk.CTkFont(weight="bold"))
        lbl_total.grid(row=0, column=5, padx=3)

        frame_btns = ctk.CTkFrame(linha, fg_color="transparent")
        frame_btns.grid(row=0, column=6, padx=3)

        ctk.CTkButton(frame_btns, text="↑", width=28, command=lambda: self.mover_item(linha, -1)).pack(side="left", padx=2)
        ctk.CTkButton(frame_btns, text="↓", width=28, command=lambda: self.mover_item(linha, 1)).pack(side="left", padx=2)
        ctk.CTkButton(frame_btns, text="X", width=28, fg_color="#DC2626", hover_color="#B91C1C",
                       command=lambda: self.remover_item_linha(linha)).pack(side="left", padx=2)

        dados = {
            "frame": linha,
            "item": lbl_item,
            "desc": ent_desc,
            "un": ent_un,
            "qtd": ent_qtd,
            "unit": ent_unit,
            "total": lbl_total,
        }

        self.linhas_itens.append(dados)

        ent_qtd.bind("<KeyRelease>", lambda e: self.atualizar_totais())
        ent_unit.bind("<KeyRelease>", lambda e: self.after_idle(lambda: self.mascara_real_time(ent_unit)))

        self.reindexar_itens()
        self.atualizar_totais()

    def remover_item_linha(self, linha_frame):
        self.linhas_itens = [l for l in self.linhas_itens if l["frame"] != linha_frame]
        linha_frame.destroy()
        self.reindexar_itens()
        self.atualizar_totais()

    def mover_item(self, linha_frame, direcao):
        idx = next(i for i, l in enumerate(self.linhas_itens) if l["frame"] == linha_frame)
        novo = idx + direcao
        if novo < 0 or novo >= len(self.linhas_itens):
            return
        self.linhas_itens[idx], self.linhas_itens[novo] = self.linhas_itens[novo], self.linhas_itens[idx]

        for l in self.linhas_itens:
            l["frame"].pack_forget()
        for l in self.linhas_itens:
            l["frame"].pack(fill="x", pady=3, padx=2)

        self.reindexar_itens()

    def reindexar_itens(self):
        for i, linha in enumerate(self.linhas_itens, start=1):
            linha["item"].configure(text=str(i))

    def atualizar_totais(self):
        total_geral = Decimal("0")
        for linha in self.linhas_itens:
            qtd = parse_moeda(linha["qtd"].get())
            unit = parse_moeda(linha["unit"].get())
            total = qtd * unit
            linha["total"].configure(text=f"R$ {formatar_brl(total)}")
            total_geral += total

        self.label_total_geral.configure(text=f"TOTAL GERAL: R$ {formatar_brl(total_geral)}")

    def validar_itens(self):
        erros = []
        for i, linha in enumerate(self.linhas_itens, start=1):
            descricao = linha["desc"].get().strip()
            valor = parse_moeda(linha["unit"].get())
            qtd = parse_moeda(linha["qtd"].get())

            if not descricao and valor == 0 and qtd == 0:
                continue
            if not descricao:
                erros.append(f"Item {i}: descrição obrigatória.")
            if qtd <= 0:
                erros.append(f"Item {i}: quantidade inválida.")
            if valor <= 0:
                erros.append(f"Item {i}: valor unitário inválido.")

        return erros

    def importar_planilha(self):
        caminho = filedialog.askopenfilename(title="Selecionar planilha", filetypes=[("Excel", "*.xlsx")])
        if not caminho:
            return

        try:
            wb = load_workbook(caminho, data_only=True)
            ws = wb.active

            linha_header = None
            for row in ws.iter_rows(min_row=1, max_row=50):
                valores = [str(c.value).strip().lower() if c.value else "" for c in row]
                if "item" in valores and ("nome" in valores or "descrição" in valores or "descricao" in valores):
                    linha_header = row
                    break

            if not linha_header:
                raise Exception("Cabeçalho não encontrado.")

            colunas = {}
            for idx, cell in enumerate(linha_header):
                v = str(cell.value).strip().lower()
                if v == "item":
                    colunas["item"] = idx
                elif v in ["nome", "descrição", "descricao"]:
                    colunas["descricao"] = idx
                elif "preço" in v or "valor" in v:
                    colunas["valor"] = idx
                elif "quantidade" in v:
                    colunas["qtd"] = idx
                elif "unidade" in v:
                    colunas["un"] = idx

            inicio = linha_header[0].row + 1

            for l in self.linhas_itens:
                l["frame"].destroy()
            self.linhas_itens.clear()

            for row in ws.iter_rows(min_row=inicio):
                desc = row[colunas["descricao"]].value
                if not desc:
                    continue

                self.adicionar_item_linha()
                linha = self.linhas_itens[-1]

                linha["desc"].insert(0, str(desc))
                linha["qtd"].delete(0, "end")
                linha["qtd"].insert(0, str(row[colunas["qtd"]].value or 0))
                linha["unit"].delete(0, "end")
                linha["unit"].insert(0, formatar_brl(Decimal(str(row[colunas["valor"]].value or 0))))
                linha["un"].delete(0, "end")
                linha["un"].insert(0, str(row[colunas["un"]].value or "UN"))

            self.atualizar_totais()
            messagebox.showinfo("Importação", "Planilha importada com sucesso.")

        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def coletar_dados(self):
        objeto = self.entry_objeto.get().strip()
        necessidade = self.text_necessidade.get("0.0", "end").strip()

        erros = self.validar_itens()
        if erros:
            messagebox.showerror("Validação", "\n".join(erros))
            return None

        lista_itens = []
        idx = 1

        for linha in self.linhas_itens:
            descricao = linha["desc"].get().strip()
            valor = parse_moeda(linha["unit"].get())

            if not descricao and valor == 0:
                continue

            lista_itens.append({
                "Item": str(idx),
                "Descrição": descricao,
                "UN": linha["un"].get(),
                "Qtd": linha["qtd"].get(),
                "Vlr Unit. (R$)": linha["unit"].get(),
                "Total": linha["total"].cget("text"),
            })
            idx += 1

        texto_itens = "__TABLE__" + json.dumps(lista_itens) if lista_itens else "[Nenhum item informado]"

        return {
            "{{OBJETO}}": objeto or "[Objeto não informado]",
            "{{NECESSIDADE}}": necessidade or "[Justificativa não informada]",
            "{{ITENS}}": texto_itens,
            "{{VALOR_ESTIMADO}}": self.label_total_geral.cget("text").replace("TOTAL GERAL: ", ""),
        }